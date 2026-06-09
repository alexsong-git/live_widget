#!/usr/bin/env python3
"""
从 Seel 接口拉取各 merchant 的 widget runMode，写回「live_widget登陆店铺.xlsx」的 MODE 列；
若 MODE 不是 PRODUCTION，则将该行 STATUS 置为 1（pytest 仅跑 STATUS=0 的店）。

用法（在项目根目录）::

    python scripts/sync_widget_modes.py

跑 UI 测试前建议先执行本脚本，再 ``pytest``。
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import openpyxl

# 与 tests/test_seel_support_icon.py 中 Excel 路径一致
DEFAULT_EXCEL = Path(__file__).resolve().parent.parent / "live_widget登陆店铺.xlsx"

API_BASE = "https://api.seel.com/gateway/ai-support-center/api/connectors/get-connector-config"
CONNECTOR_SOURCE = "WIDGET"

MERCHANT_HEADER = "MERCHANTID"
MODE_HEADER = "MODE"
STATUS_HEADER = "STATUS"

# 与浏览器请求尽量一致（接口若校验 UA/Accept 时更稳）
DEFAULT_HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US",
    "Cache-Control": "no-cache",
    "Origin": "https://seel-test-ew01.myshopify.com",
    "Referer": "https://seel-test-ew01.myshopify.com/",
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
    ),
}


def _header_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    """第 1 行：表头名（strip + casefold）-> 列号（1-based）。同名取第一次。"""
    col: dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        key = str(v).strip().casefold()
        if key and key not in col:
            col[key] = c
    return col


def _ensure_column(ws: openpyxl.worksheet.worksheet.Worksheet, col: dict[str, int], header: str) -> int:
    """若无该表头则追加一列；返回列号（1-based）。"""
    key = header.casefold()
    if key in col:
        return col[key]
    new_c = (ws.max_column or 0) + 1
    ws.cell(1, new_c).value = header
    col[key] = new_c
    return new_c


def fetch_run_mode(merchant_id: str, timeout_s: float = 60.0) -> str:
    """请求接口，返回 data.runMode 字符串。"""
    q = urlencode({"merchantId": merchant_id.strip(), "connectorSource": CONNECTOR_SOURCE})
    url = f"{API_BASE}?{q}"
    req = Request(url, headers=DEFAULT_HEADERS, method="GET")
    with urlopen(req, timeout=timeout_s) as resp:
        raw = resp.read().decode("utf-8")
    body = json.loads(raw)
    if body.get("code") != 0:
        msg = body.get("message") or str(body)
        raise RuntimeError(f"接口业务错误: {msg}")
    data = body.get("data")
    if not isinstance(data, dict):
        raise RuntimeError("响应缺少 data 对象")
    mode = data.get("runMode")
    if mode is None:
        raise RuntimeError("响应 data 中缺少 runMode")
    return str(mode).strip()


def sync_excel_modes(excel_path: Path | None = None) -> int:
    """
    读取 MERCHANTID，写入 MODE；MODE 非 PRODUCTION 或接口失败时将该行 STATUS 置为 1。
    成功返回 0；有行失败返回 1（仍会尽量写回其它行）。
    """
    path = excel_path or DEFAULT_EXCEL
    if not path.is_file():
        print(f"错误: 找不到 Excel 文件: {path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        ws = wb.active
        col = _header_columns(ws)
        mid_key = MERCHANT_HEADER.casefold()
        if mid_key not in col:
            print(
                f"错误: 表头第一行须包含 {MERCHANT_HEADER!r} 列（当前表头已扫描）。",
                file=sys.stderr,
            )
            return 1
        mid_col = col[mid_key]
        mode_col = _ensure_column(ws, col, MODE_HEADER)
        status_col = _ensure_column(ws, col, STATUS_HEADER)

        cache: dict[str, str] = {}
        errors: list[str] = []
        max_r = ws.max_row or 1

        for r in range(2, max_r + 1):
            raw_mid = ws.cell(r, mid_col).value
            if raw_mid is None or str(raw_mid).strip() == "":
                ws.cell(r, mode_col).value = None
                continue

            merchant_id = str(raw_mid).strip()
            try:
                if merchant_id in cache:
                    mode = cache[merchant_id]
                else:
                    mode = fetch_run_mode(merchant_id)
                    cache[merchant_id] = mode
                ws.cell(r, mode_col).value = mode
                line = f"行 {r}: merchantId={merchant_id} -> MODE={mode}"
                if mode.casefold() != "production":
                    ws.cell(r, status_col).value = 1
                    line += " -> STATUS=1（非 PRODUCTION）"
                print(line)
            except (HTTPError, URLError, OSError, TimeoutError, RuntimeError, json.JSONDecodeError) as e:
                err = f"行 {r} merchantId={merchant_id!r}: {e}"
                print(err, file=sys.stderr)
                errors.append(err)
                ws.cell(r, mode_col).value = f"ERROR: {e}"[:500]
                ws.cell(r, status_col).value = 1

        wb.save(path)
    finally:
        wb.close()

    if errors:
        print(
            f"\n完成但存在 {len(errors)} 处错误，已写入 MODE 为 ERROR 前缀，"
            f"对应行 {STATUS_HEADER} 已置为 1。",
            file=sys.stderr,
        )
        return 1
    print(f"\n已保存: {path}")
    return 0


def main() -> None:
    code = sync_excel_modes()
    raise SystemExit(code)


if __name__ == "__main__":
    main()
